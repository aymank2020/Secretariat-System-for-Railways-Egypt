from io import BytesIO

from openpyxl import load_workbook

from app.models.sadir import SadirModel
from app.models.warid import WaridModel

# Column mapping: (Excel header, model field, is_date)
WARID_COLUMNS = [
    ("رقم القيد", "qaid_number", False),
    ("تاريخ القيد", "qaid_date", True),
    ("جهة الوارد", "source_administration", False),
    ("رقم الكتاب", "letter_number", False),
    ("تاريخ الكتاب", "letter_date", True),
    ("رقم الوارد رئيس المجلس", "chairman_incoming_number", False),
    ("تاريخ الوارد رئيس المجلس", "chairman_incoming_date", True),
    ("رقم العودة من رئيس المجلس", "chairman_return_number", False),
    ("تاريخ العودة من رئيس المجلس", "chairman_return_date", True),
    ("عدد المرفقات", "attachment_count", False),
    ("الموضوع", "subject", False),
    ("ملاحظات", "notes", False),
    ("المستلم 1", "recipient1_name", False),
    ("تاريخ التسليم 1", "recipient1_delivery_date", True),
    ("المستلم 2", "recipient2_name", False),
    ("تاريخ التسليم 2", "recipient2_delivery_date", True),
    ("المستلم 3", "recipient3_name", False),
    ("تاريخ التسليم 3", "recipient3_delivery_date", True),
    ("وزارة", "is_ministry", False),
    ("هيئة", "is_authority", False),
    ("أخرى", "is_other", False),
    ("تفاصيل أخرى", "other_details", False),
]

SADIR_COLUMNS = [
    ("رقم القيد", "qaid_number", False),
    ("تاريخ القيد", "qaid_date", True),
    ("جهة الصادر", "destination_administration", False),
    ("رقم الكتاب", "letter_number", False),
    ("تاريخ الكتاب", "letter_date", True),
    ("رقم الوارد رئيس المجلس", "chairman_incoming_number", False),
    ("تاريخ الوارد رئيس المجلس", "chairman_incoming_date", True),
    ("رقم العودة من رئيس المجلس", "chairman_return_number", False),
    ("تاريخ العودة من رئيس المجلس", "chairman_return_date", True),
    ("عدد المرفقات", "attachment_count", False),
    ("الموضوع", "subject", False),
    ("ملاحظات", "notes", False),
    ("مرسل إليه 1", "sent_to1_name", False),
    ("تاريخ الإرسال 1", "sent_to1_delivery_date", True),
    ("مرسل إليه 2", "sent_to2_name", False),
    ("تاريخ الإرسال 2", "sent_to2_delivery_date", True),
    ("مرسل إليه 3", "sent_to3_name", False),
    ("تاريخ الإرسال 3", "sent_to3_delivery_date", True),
    ("وزارة", "is_ministry", False),
    ("هيئة", "is_authority", False),
    ("أخرى", "is_other", False),
    ("تفاصيل أخرى", "other_details", False),
]


def _parse_cell(value, is_date: bool):
    if value is None:
        return None
    if is_date:
        if hasattr(value, "isoformat"):
            return value
        # try parsing string date
        import datetime
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
            try:
                return datetime.datetime.strptime(str(value).strip(), fmt)
            except ValueError:
                continue
        return None
    # Boolean handling
    if isinstance(value, str):
        v = value.strip().lower()
        if v in ("yes", "نعم", "true", "1", "✓", "✅"):
            return True
        if v in ("no", "لا", "false", "0", "✗"):
            return False
    return value


def _parse_sheet(ws, columns_config):
    rows = []
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None for cell in row):
            continue
        record = {}
        row_errors = []
        for col_idx, (header, field, is_date) in enumerate(columns_config):
            if col_idx >= len(row):
                continue
            val = _parse_cell(row[col_idx], is_date)
            if val is None and is_date:
                # dates are optional
                continue
            record[field] = val
        if not record.get("qaid_number") or not record.get("subject"):
            errors.append({"row": row_idx, "error": "Missing required fields (qaid_number / subject)"})
            continue
        rows.append(record)
    return rows, errors


def import_warid_from_excel(db, file_bytes: bytes, file_name: str, user_id: int, user_name: str) -> dict:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    records, errors = _parse_sheet(ws, WARID_COLUMNS)

    success_count = 0
    for rec in records:
        try:
            warid = WaridModel(**rec, created_by=user_id, created_by_name=user_name)
            db.add(warid)
            db.commit()
            success_count += 1
        except Exception as e:
            db.rollback()
            errors.append({"row": "N/A", "error": str(e)})

    return {
        "success_count": success_count,
        "error_count": len(errors),
        "errors": errors,
        "total_rows": len(records) + len(errors),
    }


def import_sadir_from_excel(db, file_bytes: bytes, file_name: str, user_id: int, user_name: str) -> dict:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    records, errors = _parse_sheet(ws, SADIR_COLUMNS)

    success_count = 0
    for rec in records:
        try:
            sadir = SadirModel(**rec, created_by=user_id, created_by_name=user_name)
            db.add(sadir)
            db.commit()
            success_count += 1
        except Exception as e:
            db.rollback()
            errors.append({"row": "N/A", "error": str(e)})

    return {
        "success_count": success_count,
        "error_count": len(errors),
        "errors": errors,
        "total_rows": len(records) + len(errors),
    }
